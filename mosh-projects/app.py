# import test 
# from test import lbs_to_kg
# import utils

# # print(lbs_to_kg(150))
# print("___________________")
# # print(test.kg_to_lbs(68.02721088435374))
# numbers = [10, 3, 6, 2,11]
# print (utils.find__max(numbers))
# print (max(numbers))
# import ecommerce.shipping
# ecommerce.shipping.calc_shipping()
    # import random
    # for i in range(3):
    #     print(random.randint(0,10))


# members = ['john','mary','bob', 'naol']
# # random.choice
# print(random.choice(members))


# import random
# class Dice:
#     def roll(self):
#         first = random.randint(1,6)
#         second = random.randint(1,6)
#         return (first, second)


# dice = Dice()
# print(dice.roll())
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell1 = sheet['a1']
cell2 = sheet.cell(1,1)
# print(cell1.value)
# print(cell2.value)

# print(sheet.max_row)
        # for row in range(2, sheet.max_row + 1):
        #       cell = sheet.cell(row, 3)
        #       corrected_price = cell.value*0.9
        #       corrected_price_cell = sheet.cell(row, 4)
        #       corrected_price_cell.value = corrected_price

        # values = Reference(sheet, 
        #           min_row=2, 
        #           max_row=sheet.max_row,
        #           min_col=4,
        #           max_col=4 )
        # chart = BarChart()
        # chart.add_data(values)
        # sheet.add_chart(chart, 'e2')
        # wb.save('transactions3.xlsx')

# import matplotlib.pyplot as plt
# import networkx as nx

# # Create WBS structure
# wbs = nx.DiGraph()
# wbs.add_edges_from([
#     ("Project", "Phase 1"),
#     ("Project", "Phase 2"),
#     ("Phase 1", "Task 1.1"),
#     ("Phase 1", "Task 1.2"),
#     ("Phase 2", "Task 2.1"),
#     ("Phase 2", "Task 2.2"),
# ])

# # Draw the tree
# pos = nx.spring_layout(wbs)
# nx.draw(wbs, pos, with_labels=True, node_color='lightblue', 
#         node_size=3000, font_size=8, arrows=False)
# plt.show()